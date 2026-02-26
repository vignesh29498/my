"""
Enhanced Excel Formula Column Adder
====================================
Reads an Excel file, adds new columns with Excel formulas (VLOOKUP, IF, TEXT,
IFERROR, nested formulas, etc.), imports sheets from external Excel files,
and supports column-name-based references.

All configuration is in formulas.json (editable separately).

Usage:
    python add_formula_columns.py <input_excel_file> [output_excel_file] [config_file]

Examples:
    python add_formula_columns.py data.xlsx
    python add_formula_columns.py data.xlsx output.xlsx
    python add_formula_columns.py data.xlsx output.xlsx my_formulas.json

If no output file is specified, it creates: <input_name>_with_formulas.xlsx
If no config file is specified, it uses formulas.json from the script directory.
"""

import json
import re
import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ===========================================================================
# CONFIG LOADER
# ===========================================================================

def load_config(config_path="formulas.json"):
    """Load the full configuration from the JSON config file."""
    if not os.path.exists(config_path):
        print(f"ERROR: Config file '{config_path}' not found.")
        print("Make sure 'formulas.json' is in the same directory as this script.")
        sys.exit(1)

    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    return config


# ===========================================================================
# COLUMN HEADER -> LETTER MAPPING
# ===========================================================================

def build_column_map(ws):
    """
    Build a mapping of column header names -> column letters for a worksheet.
    Reads row 1 (header row) and maps each header to its column letter.

    Returns:
        dict: e.g. {"Issue key": "A", "Assignee": "G", ...}
    """
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header is not None:
            header_str = str(header).strip()
            col_letter = get_column_letter(col_idx)
            col_map[header_str] = col_letter
    return col_map


def resolve_formula(formula_template, row_num, col_map):
    """
    Resolve a formula template by replacing:
      - {row}          -> actual row number (e.g., 2, 3, 4...)
      - {ColumnName}   -> column letter from the header map

    The pattern {ColumnName}{row} becomes e.g. G2, G3, etc.

    Also handles column names with special chars like:
      {Custom field (Epic Link)}{row}  ->  D2

    Args:
        formula_template: Formula string with placeholders.
        row_num:          Current row number.
        col_map:          Dict mapping header names to column letters.

    Returns:
        Resolved formula string ready to insert into Excel.
    """
    formula = formula_template

    # First, find all {ColName} patterns (excluding {row})
    # Pattern: { ... } where content is NOT just "row"
    # We need to handle nested parentheses in column names like {Custom field (Epic Link)}
    placeholders = re.findall(r'\{([^}]+)\}', formula)

    for placeholder in placeholders:
        if placeholder == "row":
            continue
        # Look up the column name in the map
        if placeholder in col_map:
            col_letter = col_map[placeholder]
            # Replace {ColName} with the column letter
            formula = formula.replace("{" + placeholder + "}", col_letter)
        else:
            print(f"    WARNING: Column header '{placeholder}' not found in sheet headers. "
                  f"Available: {list(col_map.keys())}")
            # Leave it as-is so user can see what's wrong
            # But still replace with a placeholder to avoid formula errors
            formula = formula.replace("{" + placeholder + "}", "??")

    # Now replace {row} with the actual row number
    formula = formula.replace("{row}", str(row_num))

    return formula


# ===========================================================================
# IMPORT SHEETS FROM EXTERNAL FILES
# ===========================================================================

def import_external_sheets(wb, import_configs, base_dir):
    """
    Import sheets from external Excel files into the target workbook.

    Args:
        wb:              Target workbook (openpyxl Workbook).
        import_configs:  List of import configurations from JSON.
        base_dir:        Base directory for resolving relative file paths.
    """
    if not import_configs:
        print("\nNo external sheets to import.")
        return

    print(f"\n--- Importing {len(import_configs)} external sheet(s) ---")

    for imp in import_configs:
        source_file = imp["source_file"]
        source_sheet = imp.get("source_sheet", None)  # None = active sheet
        target_name = imp.get("target_sheet_name", source_sheet or "Imported")

        # Resolve relative path
        source_path = os.path.join(base_dir, source_file) if not os.path.isabs(source_file) else source_file

        if not os.path.exists(source_path):
            print(f"  WARNING: External file '{source_path}' not found. Skipping import of '{target_name}'.")
            continue

        print(f"  Importing '{source_sheet or 'active sheet'}' from '{source_file}' -> '{target_name}'")

        # Load external workbook
        ext_wb = load_workbook(source_path, data_only=False)

        if source_sheet:
            if source_sheet not in ext_wb.sheetnames:
                print(f"    WARNING: Sheet '{source_sheet}' not found in '{source_file}'. "
                      f"Available: {ext_wb.sheetnames}. Skipping.")
                continue
            ext_ws = ext_wb[source_sheet]
        else:
            ext_ws = ext_wb.active

        # Create or get target sheet in output workbook
        if target_name in wb.sheetnames:
            print(f"    Sheet '{target_name}' already exists. Overwriting data.")
            target_ws = wb[target_name]
        else:
            target_ws = wb.create_sheet(title=target_name)

        # Copy all cell values, styles not deeply copied (just values & formulas)
        for row in ext_ws.iter_rows(min_row=1, max_row=ext_ws.max_row,
                                     min_col=1, max_col=ext_ws.max_column):
            for cell in row:
                target_cell = target_ws.cell(row=cell.row, column=cell.column)
                target_cell.value = cell.value
                # Copy number format (important for dates, percentages, etc.)
                if cell.number_format:
                    target_cell.number_format = cell.number_format

        # Copy column widths
        for col_idx in range(1, ext_ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if ext_ws.column_dimensions[col_letter].width:
                target_ws.column_dimensions[col_letter].width = ext_ws.column_dimensions[col_letter].width

        print(f"    Imported {ext_ws.max_row} rows x {ext_ws.max_column} columns.")

        ext_wb.close()


# ===========================================================================
# ADD FORMULA COLUMNS TO SHEETS
# ===========================================================================

def add_formula_columns_to_sheet(wb, sheet_config):
    """
    Add new formula columns to a specific sheet in the workbook.

    Args:
        wb:            The workbook (openpyxl Workbook).
        sheet_config:  Dict with 'sheet_name' and 'new_columns' from JSON config.
    """
    sheet_name = sheet_config["sheet_name"]
    new_columns = sheet_config.get("new_columns", [])

    if not new_columns:
        print(f"\n  No new columns defined for sheet '{sheet_name}'. Skipping.")
        return

    # Get the worksheet
    if sheet_name not in wb.sheetnames:
        print(f"\n  WARNING: Sheet '{sheet_name}' not found in workbook. "
              f"Available: {wb.sheetnames}. Skipping.")
        return

    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"\n--- Processing sheet '{sheet_name}': {max_row} rows x {max_col} columns ---")

    # Build column map from existing headers (including any previously added columns)
    col_map = build_column_map(ws)
    print(f"  Column map built: {len(col_map)} headers detected.")

    # Add each new column
    for i, col_def in enumerate(new_columns):
        new_col_index = max_col + 1 + i
        new_col_letter = get_column_letter(new_col_index)
        col_name = col_def["column_name"]
        formula_template = col_def["formula"]
        description = col_def.get("description", "")

        # Write header in row 1
        ws.cell(row=1, column=new_col_index, value=col_name)

        # Update column map so subsequent formulas can reference this new column
        col_map[col_name] = new_col_letter

        # Write formula for each data row (row 2 onwards)
        for row_num in range(2, max_row + 1):
            formula = resolve_formula(formula_template, row_num, col_map)
            ws.cell(row=row_num, column=new_col_index, value=formula)

        # Show a sample resolved formula (row 2)
        sample = resolve_formula(formula_template, 2, col_map)
        print(f"  [{new_col_letter}] '{col_name}': {formula_template}")
        print(f"       -> Sample (row 2): {sample}")

    total_new = len(new_columns)
    print(f"  Added {total_new} columns to '{sheet_name}'. "
          f"Total columns now: {max_col + total_new}")


# ===========================================================================
# MAIN PROCESSING
# ===========================================================================

def process_excel(input_file, output_file=None, config_path="formulas.json"):
    """
    Main function: load config, import sheets, add formula columns, save output.

    Args:
        input_file:  Path to the input Excel file.
        output_file: Path to the output Excel file (optional).
        config_path: Path to the formulas JSON config file.
    """
    # --- Validate input ---
    if not os.path.exists(input_file):
        print(f"ERROR: Input file '{input_file}' not found.")
        sys.exit(1)

    # --- Default output file name ---
    if output_file is None:
        name, ext = os.path.splitext(input_file)
        output_file = f"{name}_with_formulas.xlsx"

    # --- Load config ---
    config = load_config(config_path)
    import_sheets = config.get("import_sheets", [])
    target_sheets = config.get("target_sheets", [])

    total_formulas = sum(len(ts.get("new_columns", [])) for ts in target_sheets)
    print(f"Config loaded: {len(import_sheets)} sheet import(s), "
          f"{len(target_sheets)} target sheet(s), {total_formulas} formula column(s).")

    # --- Load workbook ---
    print(f"\nLoading workbook '{input_file}'...")
    wb = load_workbook(input_file)
    print(f"Sheets found: {wb.sheetnames}")

    # --- Base directory for resolving relative paths ---
    base_dir = os.path.dirname(os.path.abspath(input_file))

    # --- Step 1: Import external sheets ---
    import_external_sheets(wb, import_sheets, base_dir)

    # --- Step 2: Add formula columns to each target sheet ---
    for sheet_config in target_sheets:
        add_formula_columns_to_sheet(wb, sheet_config)

    # --- Step 3: Save output ---
    wb.save(output_file)
    print(f"\n{'='*60}")
    print(f"SUCCESS! Output saved to: {output_file}")
    print(f"Sheets in output: {wb.sheetnames}")
    print(f"{'='*60}")


# ===========================================================================
# ENTRY POINT
# ===========================================================================

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    config_arg = sys.argv[3] if len(sys.argv) > 3 else None

    # Config file: use argument, or default to formulas.json in script directory
    if config_arg:
        config_file = config_arg
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_file = os.path.join(script_dir, "formulas.json")

    process_excel(input_path, output_path, config_file)
